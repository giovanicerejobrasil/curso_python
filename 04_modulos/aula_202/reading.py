"""
openpyxl - ler e alterar dados de uma planilha Excel xlsx, xlsm, xltx e xltm

Com essa biblioteca será possível ler e escrever dados em células
específicas, formatar células, inserir gráficos,
criar fórmulas, adicionar imagens e outros elementos gráficos às suas
planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
Excel, como a criação de relatórios e análise de dados e/ou facilitando a
manipulação de grandes quantidades de informações.

Instalação necessária: pip install openpyxl

Documentação: https://openpyxl.readthedocs.io/en/stable/
"""
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


def print_division(
    text='',
    sep_text='=',
    num_repeat=35,
    jump=False
):
    print() if jump else ...
    print(sep_text * num_repeat, sep='\n')
    print(text.upper(), end='\n\n')


ROOT_PATH = Path(__file__).parent
WORKBOOK_PATH = ROOT_PATH / 'workbook.xlsx'

# Carregando um arquivo excel
workbook: Workbook = load_workbook(WORKBOOK_PATH)

# Nome da planilha
sheet_name = 'My Sheet'

# Seleção da planilha
worksheet: Worksheet = workbook[sheet_name]  # type: ignore

print_division(text='sheetnames')
print(workbook.sheetnames)

print_division(text='iter_rows', jump=True)
row: tuple[Cell]
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')

        if cell.value == 'Malu':
            worksheet.cell(cell.row, 2, 3)  # type: ignore
    print()

worksheet['C3'].value = 6.5

workbook.save(WORKBOOK_PATH)
