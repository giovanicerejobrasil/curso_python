"""
openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)

Com essa biblioteca será possível ler e escrever dados em células
específicas, formatar células, inserir gráficos,
criar fórmulas, adicionar imagens e outros elementos gráficos às suas
planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
Excel, como a criação de relatórios e análise de dados e/ou facilitando a
manipulação de grandes quantidades de informações.

Instalação necessária: pip install openpyxl

Documentação: https://openpyxl.readthedocs.io/en/stable/
"""
import random
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def print_division(
    text='',
    sep_text='=',
    num_repeat=35,
    jump=False
):
    print() if jump else ...
    print(sep_text * num_repeat, sep='\n')
    print(text.upper(), end='\n\n')


ROOT_PATH = Path(__file__).parent
WORKBOOK_PATH = ROOT_PATH / 'workbook.xlsx'

workbook: Workbook = Workbook()
# planilha ativa
# worksheet: Worksheet = workbook.active

# Nome da planilha
sheet_name = 'My Sheet'

# Criação da planilha
workbook.create_sheet(sheet_name, 0)

# Seleção da planilha
worksheet: Worksheet = workbook[sheet_name]  # type: ignore

# Remover uma planilha
workbook.remove(workbook['Sheet'])  # type: ignore

print_division(text='sheetnames')
print(workbook.sheetnames)

print_division(text='cell')
# Criando os cabeçalhos
worksheet.cell(1, 1, 'Name')
worksheet.cell(1, 2, 'Age')
worksheet.cell(1, 3, 'Grade')

print('Células adicionadas com sucesso!')


def random_num():
    return round(random.uniform(0.0, 10.0), 2)


students = [
    # name, age, grade
    ['Giovani', 27, random_num()],
    ['Luanna', 24, random_num()],
    ['Mel', 11, random_num()],
    ['Malu', 2, random_num()],
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)
#         print(i, j, student_column)

print_division(text='append', jump=True)
for student in students:
    worksheet.append(student)

print('Adicionado com sucesso!')

print_division(text='save', jump=True)
workbook.save(WORKBOOK_PATH)

print('Salvo com sucesso!')
